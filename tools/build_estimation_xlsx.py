"""
Generate FoodDelivery_Estimation.xlsx
4 sheets matching template layout from BẢN MÔ TẢ YÊU CẦU SẢN PHẨM.xlsx
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

OUT = r'D:\SoLi-Food-Order-and-Deliver-App\apps\api\docs\Final_Documents\FoodDelivery_Estimation.xlsx'

# ============================================================
# FUNCTION INVENTORY (full product, not only coded scope)
# Columns per row: (Main, Sub, Web, Mobile, C1, C2, C3, C4, C5)
# Weights: W1=4, W2=5, W3=4, W4=10, W5=7
# C1=Input, C2=Output, C3=Inquiry, C4=ILF, C5=EIF
# ============================================================
W1, W2, W3, W4, W5 = 4, 5, 4, 10, 7

FUNCS = [
    # 1. Authentication & Access
    ("Authentication & Access", "Sign Up",                 "",  "X", 1, 1, 0, 1, 1),  # creates user; sends OTP via external
    ("Authentication & Access", "Sign In",                 "X", "X", 1, 1, 0, 0, 0),
    ("Authentication & Access", "Sign Out",                "X", "X", 1, 1, 0, 0, 0),
    ("Authentication & Access", "Forgot Password / Send OTP","X","X",1, 1, 0, 0, 1),  # SMS/Email gateway
    ("Authentication & Access", "Reset Password",          "X", "X", 1, 1, 0, 1, 0),
    ("Authentication & Access", "Verify Phone OTP",        "",  "X", 1, 1, 0, 1, 1),
    ("Authentication & Access", "Get Session / Refresh",   "X", "X", 0, 1, 1, 0, 0),

    # 2. User Profile & Address
    ("User Profile",            "View Profile",            "X", "X", 0, 0, 1, 0, 0),
    ("User Profile",            "Update Profile",          "X", "X", 1, 1, 0, 1, 0),
    ("User Profile",            "Upload Avatar",           "X", "X", 1, 1, 0, 0, 1),  # Cloudinary
    ("User Profile",            "Manage Delivery Addresses","",  "X", 1, 1, 1, 1, 0),

    # 3. Restaurant Discovery
    ("Restaurant Discovery",    "Browse Restaurant List",  "",  "X", 0, 0, 1, 0, 0),
    ("Restaurant Discovery",    "View Restaurant Detail",  "",  "X", 0, 0, 1, 0, 0),
    ("Restaurant Discovery",    "Search Restaurants",      "",  "X", 1, 0, 1, 0, 0),
    ("Restaurant Discovery",    "Filter / Sort Restaurants","", "X", 1, 0, 1, 0, 0),
    ("Restaurant Discovery",    "View Menu & Categories",  "",  "X", 0, 0, 1, 0, 0),
    ("Restaurant Discovery",    "View Menu Item Detail",   "",  "X", 0, 0, 1, 0, 0),
    ("Restaurant Discovery",    "Check Delivery Zone Coverage","","X",1,1, 1, 0, 0),

    # 4. Cart & Checkout
    ("Cart & Checkout",         "Add Item To Cart",        "",  "X", 1, 1, 0, 1, 0),  # cart in Redis (ILF)
    ("Cart & Checkout",         "Update Cart Item Quantity","",  "X", 1, 1, 0, 1, 0),
    ("Cart & Checkout",         "Remove Cart Item",        "",  "X", 1, 1, 0, 1, 0),
    ("Cart & Checkout",         "View Cart",               "",  "X", 0, 0, 1, 0, 0),
    ("Cart & Checkout",         "Apply Promotion / Coupon","",  "X", 1, 1, 1, 0, 0),
    ("Cart & Checkout",         "Select Delivery Address", "",  "X", 1, 0, 1, 0, 0),
    ("Cart & Checkout",         "Choose Payment Method",   "",  "X", 1, 0, 1, 0, 0),
    ("Cart & Checkout",         "Place Order (Checkout)",  "",  "X", 1, 1, 1, 1, 1),  # full flow

    # 5. Payment
    ("Payment",                 "Initiate VNPay Payment",  "",  "X", 1, 1, 0, 1, 1),
    ("Payment",                 "Process VNPay IPN",       "",  "",  1, 1, 0, 1, 1),
    ("Payment",                 "Handle VNPay Return URL", "",  "X", 0, 1, 1, 1, 1),
    ("Payment",                 "Payment Timeout Auto-Cancel","","",  0, 1, 0, 1, 0),
    ("Payment",                 "Refund (Admin)",          "X", "",  1, 1, 0, 1, 1),
    ("Payment",                 "View Payment History",    "X", "X", 0, 0, 1, 0, 0),

    # 6. Order (Customer)
    ("Order (Customer)",        "View Order History",      "",  "X", 0, 0, 1, 0, 0),
    ("Order (Customer)",        "View Order Detail",       "",  "X", 0, 0, 1, 0, 0),
    ("Order (Customer)",        "Track Order Status (Realtime)","","X",0,1,1, 0, 1),  # WebSocket
    ("Order (Customer)",        "Cancel Order",            "",  "X", 1, 1, 0, 1, 0),
    ("Order (Customer)",        "Reorder",                 "",  "X", 1, 1, 1, 1, 0),

    # 7. Notification
    ("Notification",            "Receive In-App Notification","X","X",0,1, 0, 1, 0),
    ("Notification",            "Receive Push Notification","",  "X", 0, 1, 0, 0, 1),  # FCM
    ("Notification",            "Receive Email Notification","X","X",0, 1, 0, 0, 1),  # SMTP
    ("Notification",            "View Notification History","X", "X",0, 0, 1, 0, 0),
    ("Notification",            "Update Notification Preferences","X","X",1,1,1, 1, 0),
    ("Notification",            "Register / Refresh Device Token","","X",1,1,0, 1, 1),

    # 8. Restaurant Management (Web)
    ("Restaurant Management",   "Create/Update Restaurant Profile","X","",1,1,1, 1, 0),
    ("Restaurant Management",   "Manage Menu Categories",  "X", "",  1, 1, 1, 1, 0),
    ("Restaurant Management",   "Create Menu Item",        "X", "",  1, 1, 0, 1, 1),  # image upload
    ("Restaurant Management",   "Update Menu Item",        "X", "",  1, 1, 0, 1, 1),
    ("Restaurant Management",   "Delete / Hide Menu Item", "X", "",  1, 1, 0, 1, 0),
    ("Restaurant Management",   "Manage Modifiers / Options","X","",1, 1, 1, 1, 0),
    ("Restaurant Management",   "Manage Delivery Zones",   "X", "",  1, 1, 1, 1, 0),
    ("Restaurant Management",   "Set Operating Hours",     "X", "",  1, 1, 1, 1, 0),
    ("Restaurant Management",   "View Incoming Orders (Kanban)","X","",0,1, 1, 0, 1),
    ("Restaurant Management",   "Accept / Reject Order",   "X", "",  1, 1, 0, 1, 0),
    ("Restaurant Management",   "Update Order Status (Prepare/Ready)","X","",1,1,0, 1, 0),
    ("Restaurant Management",   "Manage Restaurant Promotions","X","",1, 1, 1, 1, 0),
    ("Restaurant Management",   "View Sales Reports",      "X", "",  0, 1, 1, 0, 0),

    # 9. Shipper (Planned per ASR/SAD)
    ("Shipper Operations",      "Browse Available Deliveries","","X",0, 0, 1, 0, 0),
    ("Shipper Operations",      "Accept Delivery Assignment","",  "X",1, 1, 0, 1, 0),
    ("Shipper Operations",      "Update Pickup Status",    "",  "X", 1, 1, 0, 1, 1),
    ("Shipper Operations",      "Update Delivery Status",  "",  "X", 1, 1, 0, 1, 1),
    ("Shipper Operations",      "Live Location Update",    "",  "X", 1, 1, 0, 0, 1),
    ("Shipper Operations",      "View Earnings",           "",  "X", 0, 0, 1, 0, 0),

    # 10. Admin (Web)
    ("Admin",                   "Approve Restaurant Partner","X","",1, 1, 1, 1, 1),  # email out
    ("Admin",                   "Approve Shipper",         "X", "",  1, 1, 1, 1, 1),
    ("Admin",                   "Manage Users (List/Block)","X", "", 1, 1, 1, 1, 0),
    ("Admin",                   "Manage Platform Promotions","X","",1, 1, 1, 1, 0),
    ("Admin",                   "Monitor Live Orders",     "X", "",  0, 1, 1, 0, 1),
    ("Admin",                   "Platform Reports / Dashboards","X","",0, 1, 1, 0, 0),
    ("Admin",                   "Manage Roles & Permissions","X","",1, 1, 1, 1, 0),
    ("Admin",                   "Configure App Settings",  "X", "",  1, 1, 1, 1, 0),
    ("Admin",                   "Manage Content & Categories","X","",1,1, 1, 1, 0),
    ("Admin",                   "View Audit Logs",         "X", "",  0, 0, 1, 0, 0),

    # 11. Review & Rating (Planned)
    ("Review & Rating",         "Submit Order Review",     "",  "X", 1, 1, 0, 1, 0),
    ("Review & Rating",         "View Restaurant Reviews", "",  "X", 0, 0, 1, 0, 0),
    ("Review & Rating",         "Reply to Review (Restaurant)","X","",1,1, 1, 1, 0),
    ("Review & Rating",         "Moderate Reviews (Admin)","X", "",  1, 1, 1, 1, 0),
]

# ============================================================
# VAF Factors (12, matching template layout)
# ============================================================
VAF_FACTORS = [
    ("Bảo đảm an toàn khi cập nhật & truyền dữ liệu?", 5,
     "Thanh toán VNPay (HMAC SHA512), Better Auth bearer session, idempotency keys, HTTPS, PII hồ sơ người dùng + địa chỉ giao hàng."),
    ("Có đòi hỏi truyền thông không?", 4,
     "Mobile + Web + Backend giao tiếp REST; Socket.IO realtime cho /notifications và trạng thái đơn; webhook IPN từ VNPay; FCM push."),
    ("Có xử lý phân bố không?", 3,
     "Modular monolith + PostgreSQL + Redis/Valkey + Cloudinary + VNPay; nhiều client (mobile, web)."),
    ("Tốc độ có quan trọng không?", 4,
     "ASR ràng buộc: tìm kiếm/danh sách p95 ≤ 2s; cập nhật trạng thái ≤ 5s; xử lý đơn vào giờ cao điểm."),
    ("Có đòi hỏi cấu hình mạnh?", 3,
     "Cần DB + Redis + worker timeout ổn định; Render free tier giới hạn nên phải tối ưu kết nối, pool."),
    ("Nhập dữ liệu có transaction?", 5,
     "Place Order (CQRS) đa bước: trừ kho/khoá Redis, áp khuyến mãi, ghi đơn, gọi VNPay; transition trạng thái dùng state machine + transaction."),
    ("Dữ liệu lưu trữ cập nhật trực tuyến?", 5,
     "Trạng thái đơn, giỏ hàng, snapshot ACL, inbox thông báo đều update online; projector cập nhật ngay khi nhà hàng/menu/zone thay đổi."),
    ("Nhập/xuất/truy vấn phức tạp?", 4,
     "Tìm kiếm theo khoảng cách/zone, filter nhiều tiêu chí, lịch sử đơn theo role, báo cáo doanh thu, Kanban cho nhà hàng."),
    ("Xử lý bên trong phức tạp?", 4,
     "State machine đơn hàng, pricing engine khuyến mãi (preview/reserve/confirm/rollback), reconcile thanh toán, ACL snapshot, quiet-hours."),
    ("Mã nguồn cần thiết kế dùng lại?", 4,
     "Port/Adapter (PAYMENT_INITIATION_PORT, PROMOTION_APPLICATION_PORT), kênh thông báo (in-app/push/email) trừu tượng qua interface, EventBus."),
    ("Chuyển đổi dữ liệu & cài đặt có phức tạp?", 3,
     "Drizzle migrations + seed (catalog, zones), cấu hình CI/CD GitHub Actions + GHCR + Render IaC (Terraform)."),
    ("Dễ thay đổi & dễ dùng?", 4,
     "Người dùng cuối là khách hàng phổ thông + nhân viên nhà hàng; UI bằng tiếng Việt; mobile-first; cấu hình kênh thông báo linh hoạt."),
]

# Quick FP & VAF preview for cost sheet computation
fi_sum = sum(f[1] for f in VAF_FACTORS)
vaf = 0.65 + 0.01 * fi_sum

ufp = sum(c1*W1 + c2*W2 + c3*W3 + c4*W4 + c5*W5
          for _,_,_,_,c1,c2,c3,c4,c5 in FUNCS)
afp = round(ufp * vaf, 2)

print(f"Functions: {len(FUNCS)}")
print(f"UFP={ufp}  ΣFi={fi_sum}  VAF={vaf:.2f}  AFP={afp}")

# ============================================================
# BUILD WORKBOOK
# ============================================================
wb = openpyxl.Workbook()

# styles
thin = Side(border_style="thin", color="000000")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
hdr_fill = PatternFill("solid", fgColor="4F81BD")
sub_fill = PatternFill("solid", fgColor="D9E1F2")
total_fill = PatternFill("solid", fgColor="FFE699")
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)
hdr_font = Font(bold=True, color="FFFFFF", size=11)
bold = Font(bold=True)
title_font = Font(bold=True, size=14, color="1F4E78")

# ---------------- Sheet 1 ----------------
s1 = wb.active
s1.title = "Danh sách chức năng"

s1.merge_cells("A1:O1")
s1["A1"] = "DANH SÁCH CÁC CHỨC NĂNG"
s1["A1"].font = title_font
s1["A1"].alignment = center

s1["A2"] = "Tên dự án:"; s1["B2"] = "SoLi Food Delivery Platform"
s1["A3"] = "Người thực hiện:"; s1["B3"] = "Nhóm phát triển SoLi"
for r in (2,3):
    s1[f"A{r}"].font = bold

# header rows 5 & 6
s1.merge_cells("A5:A6"); s1["A5"] = "STT"
s1.merge_cells("B5:B6"); s1["B5"] = "Tên chức năng chính"
s1.merge_cells("C5:C6"); s1["C5"] = "Tên chức năng con"
s1.merge_cells("D5:E5"); s1["D5"] = "Môi trường triển khai"
s1.merge_cells("F5:O5"); s1["F5"] = "DANH SÁCH ĐỘ ĐO"
s1["D6"] = "Web"; s1["E6"] = "Mobile"
sub_headers = ["C1","W1","C2","W2","C3","W3","C4","W4","C5","W5"]
for i, h in enumerate(sub_headers):
    s1.cell(6, 6+i, h)

for r in (5,6):
    for c in range(1,16):
        cell = s1.cell(r,c)
        cell.font = hdr_font
        cell.alignment = center
        cell.fill = hdr_fill
        cell.border = border

# C1/C2/.. legend row above? Use note row 4
s1["F4"] = "Nhập dữ liệu";   s1["H4"] = "Xuất dữ liệu"
s1["J4"] = "Truy vấn";       s1["L4"] = "File nội bộ"; s1["N4"] = "Giao diện ngoài"
for col in ("F4","H4","J4","L4","N4"):
    s1[col].font = Font(italic=True, size=9, color="666666")
    s1[col].alignment = center

# data rows starting at row 7
start = 7
for i, (main, sub, web, mob, c1,c2,c3,c4,c5) in enumerate(FUNCS):
    r = start + i
    s1.cell(r, 1, i+1).alignment = center
    s1.cell(r, 2, main).alignment = left
    s1.cell(r, 3, sub).alignment = left
    s1.cell(r, 4, web).alignment = center
    s1.cell(r, 5, mob).alignment = center
    s1.cell(r, 6, c1).alignment = center
    s1.cell(r, 7, W1).alignment = center
    s1.cell(r, 8, c2).alignment = center
    s1.cell(r, 9, W2).alignment = center
    s1.cell(r,10, c3).alignment = center
    s1.cell(r,11, W3).alignment = center
    s1.cell(r,12, c4).alignment = center
    s1.cell(r,13, W4).alignment = center
    s1.cell(r,14, c5).alignment = center
    s1.cell(r,15, W5).alignment = center
    for c in range(1,16):
        s1.cell(r,c).border = border

last = start + len(FUNCS) - 1

# Totals row (UFP contributions per category) -> per template style:
# row TOTAL: sum of Ci*Wi
total_row = last + 1
s1.cell(total_row, 2, "Tổng").font = bold
s1.cell(total_row, 2).fill = total_fill
# Per template: place sum-of-(Ci*Wi) under each W column
s1.cell(total_row, 7,  f"=SUMPRODUCT(F{start}:F{last},G{start}:G{last})")
s1.cell(total_row, 9,  f"=SUMPRODUCT(H{start}:H{last},I{start}:I{last})")
s1.cell(total_row, 11, f"=SUMPRODUCT(J{start}:J{last},K{start}:K{last})")
s1.cell(total_row, 13, f"=SUMPRODUCT(L{start}:L{last},M{start}:M{last})")
s1.cell(total_row, 15, f"=SUMPRODUCT(N{start}:N{last},O{start}:O{last})")
# UFP grand total
s1.cell(total_row, 1, "UFP").font = bold
for c in (1,2,7,9,11,13,15):
    s1.cell(total_row, c).fill = total_fill
    s1.cell(total_row, c).font = bold
    s1.cell(total_row, c).border = border
    s1.cell(total_row, c).alignment = center

# Summary block below
summary_row = total_row + 2
s1.cell(summary_row,   2, "Tổng UFP (Unadjusted FP)").font = bold
s1.cell(summary_row,   3, f"=G{total_row}+I{total_row}+K{total_row}+M{total_row}+O{total_row}")
s1.cell(summary_row+1, 2, "ΣFi (Tổng điểm 12 yếu tố)").font = bold
s1.cell(summary_row+1, 3, "='Hệ số hiệu chỉnh'!C14")
s1.cell(summary_row+2, 2, "VAF = 0.65 + 0.01 × ΣFi").font = bold
s1.cell(summary_row+2, 3, f"=0.65+0.01*C{summary_row+1}")
s1.cell(summary_row+3, 2, "AFP = UFP × VAF (Số lượng FP)").font = bold
s1.cell(summary_row+3, 3, f"=C{summary_row}*C{summary_row+2}")
for k in range(4):
    s1.cell(summary_row+k, 2).fill = sub_fill
    s1.cell(summary_row+k, 3).fill = sub_fill
    s1.cell(summary_row+k, 2).border = border
    s1.cell(summary_row+k, 3).border = border

# Cost preview block (mirrors template rows 41-47)
cost_start = summary_row + 6
s1.cell(cost_start,   2, "Tên").font = bold
s1.cell(cost_start,   3, "Giá trị").font = bold
s1.cell(cost_start,   4, "Đơn vị").font = bold
items = [
    ("Chi phí trả mỗi người/tháng", 1500, "USD"),
    ("Năng suất giả định",             8, "FP/pm"),
]
for i, (n,v,u) in enumerate(items):
    s1.cell(cost_start+1+i, 2, n)
    s1.cell(cost_start+1+i, 3, v)
    s1.cell(cost_start+1+i, 4, u)
s1.cell(cost_start+3, 2, "Chi phí mỗi FP")
s1.cell(cost_start+3, 3, f"=C{cost_start+1}/C{cost_start+2}")
s1.cell(cost_start+3, 4, "USD/FP")
s1.cell(cost_start+4, 2, "Số lượng FP")
s1.cell(cost_start+4, 3, f"=C{summary_row+3}")
s1.cell(cost_start+4, 4, "FP")
s1.cell(cost_start+5, 2, "Quy ra đơn vị pm")
s1.cell(cost_start+5, 3, f"=C{cost_start+4}/C{cost_start+2}")
s1.cell(cost_start+5, 4, "pm")
s1.cell(cost_start+6, 2, "Chi phí quy ra tiền USD")
s1.cell(cost_start+6, 3, f"=C{cost_start+4}*C{cost_start+3}")
s1.cell(cost_start+6, 4, "USD")

for r in range(cost_start, cost_start+7):
    for c in (2,3,4):
        cell = s1.cell(r,c)
        cell.border = border
        if r == cost_start:
            cell.fill = hdr_fill; cell.font = hdr_font; cell.alignment = center
        else:
            cell.alignment = left if c==2 else center
s1.cell(cost_start+6, 2).font = bold
s1.cell(cost_start+6, 3).font = bold
s1.cell(cost_start+6, 3).fill = total_fill

# column widths
widths = [6, 26, 36, 8, 8, 6, 6, 6, 6, 6, 6, 6, 6, 6, 6]
for i, w in enumerate(widths):
    s1.column_dimensions[get_column_letter(i+1)].width = w
s1.row_dimensions[1].height = 24
s1.row_dimensions[5].height = 26
s1.row_dimensions[6].height = 20
s1.freeze_panes = "A7"

# ---------------- Sheet 2: Hệ số hiệu chỉnh ----------------
s2 = wb.create_sheet("Hệ số hiệu chỉnh")
headers2 = ["STT", "Câu hỏi (Yếu tố kỹ thuật)", "Điểm (Fi)", "Lý do chọn"]
for i, h in enumerate(headers2):
    c = s2.cell(1, i+1, h)
    c.font = hdr_font; c.fill = hdr_fill; c.alignment = center; c.border = border

for i, (q, score, reason) in enumerate(VAF_FACTORS):
    r = 2 + i
    s2.cell(r, 1, i+1).alignment = center
    s2.cell(r, 2, q).alignment = left
    s2.cell(r, 3, score).alignment = center
    s2.cell(r, 4, reason).alignment = left
    for c in range(1, 5):
        s2.cell(r, c).border = border

total2 = 2 + len(VAF_FACTORS)
s2.cell(total2, 2, "Tổng ΣFi").font = bold
s2.cell(total2, 3, f"=SUM(C2:C{total2-1})").font = bold
s2.cell(total2, 2).fill = total_fill
s2.cell(total2, 3).fill = total_fill
for c in range(1, 5):
    s2.cell(total2, c).border = border
s2.cell(total2+1, 2, "VAF = 0.65 + 0.01 × ΣFi").font = bold
s2.cell(total2+1, 3, f"=0.65+0.01*C{total2}").font = bold
s2.cell(total2+1, 2).fill = sub_fill
s2.cell(total2+1, 3).fill = sub_fill
for c in range(1, 5):
    s2.cell(total2+1, c).border = border

s2.column_dimensions["A"].width = 6
s2.column_dimensions["B"].width = 52
s2.column_dimensions["C"].width = 12
s2.column_dimensions["D"].width = 70
for r in range(2, total2):
    s2.row_dimensions[r].height = 42

# ---------------- Sheet 3: Công nghệ triển khai ----------------
s3 = wb.create_sheet("Công nghệ triển khai")
s3.merge_cells("A1:D1")
s3["A1"] = "CÔNG NGHỆ TRIỂN KHAI - SoLi Food Delivery Platform"
s3["A1"].font = title_font; s3["A1"].alignment = center

tech_headers = ["STT", "Hạng mục", "Công nghệ", "Ghi chú"]
for i, h in enumerate(tech_headers):
    c = s3.cell(3, i+1, h)
    c.font = hdr_font; c.fill = hdr_fill; c.alignment = center; c.border = border

TECH = [
    ("Frontend Web",       "React 18 + Vite + TypeScript + TailwindCSS + shadcn/ui",
     "Dashboard nhà hàng và quản trị; build production ra Nginx Docker image."),
    ("Frontend Mobile",    "React Native (Expo) + NativeWind + Expo Router",
     "Ứng dụng khách hàng; EAS Build phát hành Android. iOS sẵn sàng."),
    ("Backend",            "NestJS 10 + TypeScript (Node.js LTS)",
     "Modular monolith; module bounded contexts: auth, restaurant-catalog, ordering, payment, promotion, notification, image."),
    ("Kiến trúc",          "Modular Monolith + Selective CQRS + EventBus nội bộ",
     "PlaceOrderCommand, TransitionOrderCommand, ProcessIpnCommand; in-process EventBus + handler."),
    ("ACL / Anti-Corruption","Snapshot Projector + Repository",
     "menu-item, restaurant, delivery-zone snapshots cho Ordering; restaurant snapshot cho Notification."),
    ("Cơ sở dữ liệu",      "PostgreSQL 16 + Drizzle ORM",
     "Migration bằng drizzle-kit; schema khai báo TypeScript-first."),
    ("Cache / Realtime State","Redis / Valkey",
     "Giỏ hàng, idempotency keys, distributed lock, presence, rate limit."),
    ("Realtime Push",      "Socket.IO (/notifications namespace)",
     "Cập nhật trạng thái đơn, thông báo trong app, theo dõi shipper."),
    ("Xác thực",           "Better Auth (bearer session) + Admin plugin + Phone OTP",
     "Vai trò: admin, restaurant, shipper, user; OTP qua nhà cung cấp SMS."),
    ("Thanh toán",         "VNPay (URL build + IPN + Return URL) — HMAC SHA512",
     "Refund stub cho dev; sẵn sàng tích hợp endpoint refund thật."),
    ("Lưu trữ ảnh",        "Cloudinary (signed upload)",
     "Hình ảnh menu, ảnh đại diện; URL chuẩn hoá lưu trong DB."),
    ("Push Notification",  "Firebase Cloud Messaging (FCM v1 API)",
     "Có Stub provider cho môi trường dev/test."),
    ("Email",              "Nodemailer + SMTP",
     "Có Noop provider để test offline; template HTML cho thông báo đơn hàng."),
    ("Định vị / Bản đồ",   "Tính khoảng cách theo công thức Haversine + zone polygon",
     "Module geo.service nội bộ; không phụ thuộc Google Maps SDK ở backend."),
    ("Job / Task nền",     "NestJS Schedule + Cron",
     "order-timeout, payment-timeout, device-token-cleanup."),
    ("Validation",         "Zod + class-validator + DTO",
     "Schema env (env.schema.ts), DTO toàn module, validator VND amount."),
    ("Testing",            "Jest + Supertest + Testing Library + Playwright",
     "Unit + e2e backend; integration cho web; smoke test mobile."),
    ("CI/CD",              "GitHub Actions + GHCR + Render IaC (Terraform)",
     "pipeline-main, pipeline-api, pipeline-web, pipeline-mobile, cd-render-iac."),
    ("Đóng gói triển khai","Docker (multi-stage) + docker-compose + Nginx",
     "Image API và Web đẩy lên ghcr.io; Render tier free khi phát triển."),
    ("Logging / Quan sát", "NestJS Logger + structured logs + health endpoint",
     "Sẵn sàng tích hợp OpenTelemetry; metrics qua /health, /readyz."),
    ("Bảo mật",            "HTTPS/TLS, Helmet, rate limit, idempotency, HMAC chữ ký",
     "Tuân thủ OWASP Top 10; quản lý secret qua biến môi trường."),
    ("Quản lý phụ thuộc",  "pnpm workspaces + Turborepo",
     "Monorepo apps/api, apps/web, apps/mobile, infra/render, tools."),
]
for i, (cat, tech, note) in enumerate(TECH):
    r = 4 + i
    s3.cell(r, 1, i+1).alignment = center
    s3.cell(r, 2, cat).alignment = left
    s3.cell(r, 3, tech).alignment = left
    s3.cell(r, 4, note).alignment = left
    for c in range(1, 5):
        s3.cell(r, c).border = border
    if i % 2 == 0:
        for c in range(1, 5):
            s3.cell(r, c).fill = sub_fill

s3.column_dimensions["A"].width = 6
s3.column_dimensions["B"].width = 28
s3.column_dimensions["C"].width = 50
s3.column_dimensions["D"].width = 70
s3.row_dimensions[1].height = 26

# ---------------- Sheet 4: Danh mục chi phí ----------------
s4 = wb.create_sheet("Danh mục chi phí")
s4.merge_cells("A1:D1")
s4["A1"] = "TỔNG CHI PHÍ PHÁT TRIỂN PHẦN MỀM - SoLi Food Delivery"
s4["A1"].font = title_font; s4["A1"].alignment = center

# Bảng FP & chi phí cơ sở (referenced from Sheet 1)
s4["A3"] = "Tham số đầu vào"
s4["A3"].font = bold; s4["A3"].fill = hdr_fill; s4["A3"].font = hdr_font
s4.merge_cells("A3:D3"); s4["A3"].alignment = center
base = [
    ("Tổng UFP",                       f"='Danh sách chức năng'!C{summary_row}",      "FP"),
    ("ΣFi",                            f"='Hệ số hiệu chỉnh'!C{total2}",             ""),
    ("VAF",                            f"='Hệ số hiệu chỉnh'!C{total2+1}",           ""),
    ("AFP (FP đã hiệu chỉnh)",         f"='Danh sách chức năng'!C{summary_row+3}",    "FP"),
    ("Chi phí trả mỗi người/tháng",    1500,                                          "USD"),
    ("Năng suất giả định",             8,                                              "FP/pm"),
    ("Chi phí mỗi FP",                 "=C8/C9",                                       "USD/FP"),
    ("Quy ra đơn vị pm",               "=C7/C9",                                       "pm"),
    ("Tổng chi phí dự án (gross)",     "=C7*C10",                                      "USD"),
]
for i, (n, v, u) in enumerate(base):
    r = 4 + i
    s4.cell(r, 1, n).alignment = left
    s4.cell(r, 2).alignment = center
    s4.cell(r, 3, v).alignment = center
    s4.cell(r, 4, u).alignment = center
    for c in range(1, 5):
        s4.cell(r, c).border = border

GROSS_ROW = 4 + len(base) - 1   # row of "Tổng chi phí dự án (gross)"

# Allocation table
alloc_start = GROSS_ROW + 3
s4.cell(alloc_start, 1, "Danh mục chi phí").font = hdr_font
s4.cell(alloc_start, 2, "Tỷ lệ").font = hdr_font
s4.cell(alloc_start, 3, "Thành tiền (USD)").font = hdr_font
s4.cell(alloc_start, 4, "Diễn giải").font = hdr_font
for c in range(1,5):
    s4.cell(alloc_start, c).fill = hdr_fill
    s4.cell(alloc_start, c).alignment = center
    s4.cell(alloc_start, c).border = border

ALLOC = [
    ("Chi phí BA (Business Analysis)",      0.08, "Khảo sát, UC, BRD, business rules, user stories."),
    ("Chi phí Architecture & Design",       0.10, "ASR, ADD, ADR, SAD; thiết kế module/CQRS/ACL; bản vẽ PlantUML."),
    ("Chi phí Development",                 0.45, "Backend NestJS, Web React, Mobile React Native, tích hợp VNPay/FCM/Cloudinary."),
    ("Chi phí Testing & QA",                0.15, "Unit, integration, e2e, kịch bản thanh toán/IPN, hồi quy."),
    ("Chi phí Deployment & DevOps",         0.05, "Docker, GHCR, Render IaC, GitHub Actions pipelines."),
    ("Chi phí Training & Hỗ trợ",           0.03, "Tài liệu sử dụng cho nhà hàng, video hướng dẫn ≤5 phút."),
    ("Chi phí Documentation",               0.07, "Tài liệu kỹ thuật, sổ tay vận hành, tài liệu API."),
    ("Chi phí Dự phòng (Reserve)",          0.07, "Rủi ro tích hợp, thay đổi yêu cầu, mở rộng giai đoạn 2."),
]
for i, (name, pct, note) in enumerate(ALLOC):
    r = alloc_start + 1 + i
    s4.cell(r, 1, name).alignment = left
    s4.cell(r, 2, pct).alignment = center
    s4.cell(r, 2).number_format = "0.00%"
    s4.cell(r, 3, f"=B{r}*C{GROSS_ROW}").alignment = center
    s4.cell(r, 3).number_format = "#,##0.00"
    s4.cell(r, 4, note).alignment = left
    for c in range(1, 5):
        s4.cell(r, c).border = border

alloc_end = alloc_start + len(ALLOC)
tot = alloc_end + 1
s4.cell(tot, 1, "TỔNG CỘNG").font = bold
s4.cell(tot, 2, f"=SUM(B{alloc_start+1}:B{alloc_end})")
s4.cell(tot, 2).number_format = "0.00%"
s4.cell(tot, 3, f"=SUM(C{alloc_start+1}:C{alloc_end})")
s4.cell(tot, 3).number_format = "#,##0.00"
s4.cell(tot, 3).font = bold
for c in range(1, 5):
    s4.cell(tot, c).fill = total_fill
    s4.cell(tot, c).border = border
    s4.cell(tot, c).alignment = center
s4.cell(tot, 1).alignment = left

s4.column_dimensions["A"].width = 34
s4.column_dimensions["B"].width = 12
s4.column_dimensions["C"].width = 22
s4.column_dimensions["D"].width = 70
s4.row_dimensions[1].height = 26

wb.save(OUT)
print("Saved:", OUT)
